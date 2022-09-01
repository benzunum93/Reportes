from logging import FileHandler
import sqlite3
import pandas as pd
from datetime import datetime
import xlsxwriter
from openpyxl import load_workbook
#To move files to another directory
import shutil
import os
#Work with excel.xlsx files



def conteo_personas(personas,Fuerza):
    name_reporte='Reporte_conteo.xlsx'
    try:

        workbook= xlsxwriter.Workbook(name_reporte)

        # Formato de titulo
        formato_titulo = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '#333f4f',
            'font_color': 'white',
            'text_wrap': True})
        # Formato de variables
        formato_variables = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'left',
            'valign': 'top',
            'fg_color': '#ddebf7',
            'font_color': 'black',
            'text_wrap': True})
        # Formato del texto normal
        formato_normal = workbook.add_format({
            'border': 1,
            'align': 'left',
            'valign': 'top',
            'text_wrap': True})
        # Add a bold format to use to highlight cells.
        bold = workbook.add_format({'bold': 1})
        
        
        Unidad=str(Fuerza).capitalize()
        worksheet = workbook.add_worksheet(Unidad)
        worksheet.write(1,1, 'Tripulaciones',formato_variables)
        #Write in multiples rows, and it used for the title
        worksheet.merge_range('A1:B1', Unidad,formato_titulo)
        i,j=2,1 #Fila=i, Columna=j
        for per in personas:
            worksheet.set_column(i, 0, 15)
            worksheet.write(i,j, per)
            i+=1
        worksheet.write(i, 0, 'Total',bold)
        worksheet.write(i,j, len(personas))

        workbook.close()
        
    except FileNotFoundError as e:
        print('Por favor cierre el archivo ',name_reporte)

def horas_pilotos(nombre,horas):
    name_reporte='Reporte_Horas.xlsx'
    try:
        workbook= xlsxwriter.Workbook(name_reporte)
        # Formato de titulo
        formato_titulo = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '#333f4f',
            'font_color': 'white',
            'text_wrap': True})
        # Formato de variables

        name_pilot=str(nombre).capitalize()
        if 'Horas' in workbook.sheetnames:
            #print('Horas exists')
            worksheet = workbook.add_worksheet('Horas_2')
        else:
            worksheet = workbook.add_worksheet('Horas')
        #Create a title with the variable 'nombre'
        worksheet.merge_range('A1:B1', name_pilot,formato_titulo)
        i,j=1,0 #Fila=i, Columna=j
        
        for name in horas:
            
            worksheet.set_column(i, 0, 15)
            worksheet.write(i,j, name)
            worksheet.write(i,j+1, horas[name])
            i+=1

        workbook.close()
        
    except FileNotFoundError as e:
        print('Por favor cierre el archivo ',name_reporte)

def reporte_ejc(datos):
    
    name_reporte='Reporte del Mes EJC.xlsx'

    try:
        
        #If name_reporte doesn't exist
        
        workbook= xlsxwriter.Workbook(name_reporte)
        worksheet = workbook.add_worksheet('EJC')
        workbook.close()
        
        df=datos.iloc[:,0:16]
        
        with pd.ExcelWriter(name_reporte, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name="EJC", index=False)
        
        
        
        
        
        
    except FileNotFoundError as e:

        print('Por favor cierre el archivo ',name_reporte)



def reporte_fac(datos):
    name_reporte='Reporte del Mes FAC-FAB.xlsx'
    try:
        workbook= xlsxwriter.Workbook(name_reporte)
        
        formato_normal = workbook.add_format({
            'border': 1,
            'align': 'left',
            'valign': 'top',
            'text_wrap': True})
        
        worksheet = workbook.add_worksheet('FACFAB')
        #i,j=5,1 #Fila=i, Columna=j
        #worksheet.set_column(i, 0, 15)
        workbook.close()
        datos=datos.iloc[:,0:16]
        with pd.ExcelWriter(name_reporte, engine="openpyxl", mode="w") as writer:
            datos.to_excel(writer, sheet_name="FACFAB", index=False)
            #print('Writing data...')
        
       

        
    except FileNotFoundError as e:
        print('Por favor cierre el archivo ',name_reporte)

def move_files(name_file):
    
    file_source = 'D:\Sebastian\Course python\Django\Ciac_web_page\ciac_venv\ciac_set'
    file_destination = 'D:\Sebastian\Course python'
 
    get_files = os.listdir(file_source)
 

    shutil.move(file_source + name_file, file_destination)